"""UDPE Excel 模板导入器: xlsx → grid schemaJson.

设计文档: plans/udpe-design/outputs/m3-stage4-design.md §四

核心能力:
- 解析合并单元格 → grid row cells + span
- 占位符识别 (含 {{ xxx | filter }})
- 样式映射 (bold / fontSize / align / valign / color)
- 列数推导 (各行 cell span 之和最大值)
- 行高换算 (point → mm)
- 多 sheet 扫描

V1 限制:
- 不支持纵向合并 (Excel row_span) - UDPE grid V1 只支持横向
- 不支持公式 (用户需预计算)
- 不支持跨 sheet 引用
- 渐变/图案填充丢失
"""
import io
import logging
import re
from typing import Any, Optional

logger = logging.getLogger(__name__)

# 占位符正则: {{ path | filter1 | filter2 }}
PLACEHOLDER_RE = re.compile(r"\{\{\s*([^}]+?)\s*\}\}")


def _clean_template_text(text: str) -> str:
    """规范化占位符: {{x}} → {{ x }}, {{x|filter}} → {{ x | filter }}."""
    def repl(m: re.Match) -> str:
        inner = m.group(1).strip()
        return "{{ " + inner + " }}"
    return PLACEHOLDER_RE.sub(repl, text)


def _has_border(cell) -> bool:
    """任一边有边框 (且不是 'none') → 算有边框."""
    b = cell.border
    for side in [b.top, b.bottom, b.left, b.right]:
        if side and getattr(side, "style", None) and side.style != "none":
            return True
    return False


def _color_to_hex(color_obj) -> Optional[str]:
    """openpyxl Color 对象 → hex 字符串 (#RRGGBB)."""
    if not color_obj:
        return None
    try:
        rgb = getattr(color_obj, "rgb", None)
        if rgb and isinstance(rgb, str):
            # openpyxl 格式: '00FF0000' (AARRGGBB) 或 'FFFF0000'
            if len(rgb) == 8:
                return "#" + rgb[2:]
            if len(rgb) == 6:
                return "#" + rgb
        # theme/tint 类型暂不支持
    except Exception:
        pass
    return None


def _convert_sheet(sheet) -> dict:
    """单个 sheet → grid block schemaJson."""
    if not sheet or sheet.max_row == 0 or sheet.max_column == 0:
        return {
            "type": "grid",
            "colCount": 1,
            "border": True,
            "rows": [],
        }

    # Step 1: 扫描所有合并区域, 记录起始位置
    # 格式: cells[(r, c)] = { text, span, row_span, bold, fontSize, align, valign, color, has_border }
    cells = {}
    occupied: list[list[bool]] = [[False] * sheet.max_column for _ in range(sheet.max_row)]

    # 1.1 合并区域
    for mr in sheet.merged_cells.ranges:
        start_cell = sheet.cell(mr.min_row, mr.min_col)
        col_span = mr.max_col - mr.min_col + 1
        row_span = mr.max_row - mr.min_row + 1
        # V1: 纵向合并警告 + 降级为按行展开
        # 如果 row_span > 1: 取首行, 其他行也各放一个 cell (避免内容丢失)
        # 但 UDPE grid V1 严格说 cell 不能跨行. 这里简化为: 只在第一行放一个 cell (内容用首行)
        # 其他行也放一个空 cell 占位, 但 span=1
        if row_span > 1:
            logger.warning(
                f"[UDPE] Excel 合并区域 {mr} 纵向跨 {row_span} 行, V1 降级为单行 + 占位."
            )
        # 起始 cell (纵向合并的 _origin_row 是 min_row)
        cells[(mr.min_row, mr.min_col)] = {
            "_col": mr.min_col,
            "text": str(start_cell.value) if start_cell.value is not None else "",
            "span": col_span,
            "row_span": row_span,
            "_origin_row": mr.min_row,
            "bold": bool(start_cell.font.bold) if start_cell.font else False,
            "fontSize": int(start_cell.font.size) if start_cell.font and start_cell.font.size else None,
            "align": start_cell.alignment.horizontal if start_cell.alignment else None,
            "valign": start_cell.alignment.vertical if start_cell.alignment else None,
            "color": _color_to_hex(start_cell.font.color) if start_cell.font else None,
            "has_border": _has_border(start_cell),
        }
        # 标记占用 (含跨行)
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                occupied[r - 1][c - 1] = True

    # 1.2 非合并区域
    for row in sheet.iter_rows():
        for cell in row:
            if (cell.row, cell.column) in cells:
                continue  # 合并区域已处理
            if cell.value is None and not _has_border(cell):
                continue  # 完全空 cell
            # 跳过纯空 cell (无值无边框)
            text = str(cell.value) if cell.value is not None else ""
            cells[(cell.row, cell.column)] = {
                "_col": cell.column,
                "text": text,
                "span": 1,
                "row_span": 1,
                "_origin_row": cell.row,
                "bold": bool(cell.font.bold) if cell.font else False,
                "fontSize": int(cell.font.size) if cell.font and cell.font.size else None,
                "align": cell.alignment.horizontal if cell.alignment else None,
                "valign": cell.alignment.vertical if cell.alignment else None,
                "color": _color_to_hex(cell.font.color) if cell.font else None,
                "has_border": _has_border(cell),
            }
            occupied[cell.row - 1][cell.column - 1] = True

    # Step 2: 按行分组, 构造 grid rows
    # 关键: 纵向合并 (row_span > 1) 只在 min_row 行放 cell, 高度累加
    # 用 _origin_row 字段标记 cell 的"实际显示行", 渲染时只在 _origin_row 行放
    grid_rows: list[dict] = []
    skipped_rows: set = set()
    for r in range(1, sheet.max_row + 1):
        if r in skipped_rows:
            continue
        row_cells_out = []
        for c in range(1, sheet.max_column + 1):
            if not occupied[r - 1][c - 1]:
                continue
            if (r, c) not in cells:
                continue
            cell_meta = cells[(r, c)]
            # 只在 _origin_row 行放
            if cell_meta.get("_origin_row", r) != r:
                continue
            row_cells_out.append({
                "_col": cell_meta.get("_col", c),
                "text": _clean_template_text(cell_meta["text"]),
                "span": cell_meta["span"],
                "bold": cell_meta["bold"],
                **({"align": cell_meta["align"]} if cell_meta["align"] else {}),
                **({"valign": cell_meta["valign"]} if cell_meta["valign"] else {}),
                **({"fontSize": cell_meta["fontSize"]} if cell_meta["fontSize"] else {}),
                **({"color": cell_meta["color"]} if cell_meta["color"] else {}),
            })
        # 填充列间隙: 空列用空 cell 占位, 保持列对齐
        if row_cells_out:
            filled = []
            current_col = 1
            for cell_out in row_cells_out:
                cell_col = cell_out.pop("_col")  # 移除临时字段
                while current_col < cell_col:
                    # 插入空 cell 占位
                    filled.append({"text": "", "span": 1, "bold": False})
                    current_col += 1
                filled.append(cell_out)
                current_col += cell_out.get("span", 1)
            row_cells_out = filled
        if row_cells_out:
            # 行高: 此行 + 后续被合并的行 (row_span > 1)
            max_row_span = 1
            for (rr, cc), v in cells.items():
                if v.get("_origin_row") == r and v.get("row_span", 1) > max_row_span:
                    max_row_span = v["row_span"]
            total_height_mm = 0
            for rr_offset in range(max_row_span):
                rr = r + rr_offset
                if rr_offset > 0:
                    skipped_rows.add(rr)
                try:
                    h = sheet.row_dimensions[rr].height
                    if h:
                        total_height_mm += max(6, int(h * 0.353))
                    else:
                        total_height_mm += 14
                except Exception:
                    total_height_mm += 14
            height_mm = max(6, total_height_mm)
            grid_rows.append({"height": height_mm, "cells": row_cells_out})

    # Step 3: 推导 colCount (各行 cell span 之和的最大值)
    col_count = 1
    for row in grid_rows:
        total = sum(c.get("span", 1) for c in row["cells"])
        col_count = max(col_count, total)

    # Step 4: 边框 - V1 默认开 (用户 Excel 通常希望看到表格线)
    # 但尊重 Excel 实际边框: 如果任何 cell 有边框 → 开, 否则看 row 数 (>=2 视为表格)
    has_any_border = any(c.get("has_border", False) for r in grid_rows for c in r["cells"])
    show_border = has_any_border or len(grid_rows) >= 2

    return {
        "type": "grid",
        "colCount": col_count,
        "border": show_border,
        "borderColor": "#000000",
        "borderWidth": 1,
        "rows": grid_rows,
    }


def parse_excel(file_bytes: bytes) -> dict:
    """主入口: xlsx 字节流 → 解析所有 sheet.

    Returns:
        {
            "sheets": [
                {
                    "name": "Sheet1",
                    "rowCount": 36,
                    "colCount": 7,
                    "mergedCount": 24,
                    "schemaJson": { "body": [{ "type": "grid", ... }] },
                    "placeholders": ["form.title", "form.totalAmount", ...],
                },
                ...
            ],
            "totalSheets": 3,
        }
    """
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
    sheets_out = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        grid = _convert_sheet(ws)
        # 提取占位符
        placeholders = set()
        for row in grid.get("rows", []):
            for c in row.get("cells", []):
                txt = c.get("text", "")
                for m in PLACEHOLDER_RE.finditer(txt):
                    # 提取路径部分 (| 之前)
                    path = m.group(1).split("|")[0].strip()
                    if path:
                        placeholders.add(path)
        sheets_out.append({
            "name": sheet_name,
            "rowCount": ws.max_row or 0,
            "colCount": ws.max_column or 0,
            "mergedCount": len(ws.merged_cells.ranges),
            "schemaJson": {"body": [grid]},
            "placeholders": sorted(placeholders),
        })
    return {
        "sheets": sheets_out,
        "totalSheets": len(wb.sheetnames),
    }

"""
报销中心 Service
- 列表 / 详情 / 创建（从销售费用生成） / 编辑 / 删除
- 回填实际报销 → 自动回写销售费用
- 模板列表 / AI 生成说明 / AI 风险检测
"""
import re
import uuid
from datetime import date, datetime
from decimal import Decimal
from typing import Optional

from sqlalchemy import and_, or_, select, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.exceptions import NotFoundException, ParamErrorException
from app.modules.reimbursement.models import ReimbursementForm, ReimbursementDetail, ReimbursementTemplate
from app.modules.reimbursement.schemas import (
    ReimburseCreate, ReimburseUpdate, ReimburseFillback, ReimburseDelete,
)
from app.modules.reimbursement.templates_builtin import (
    get_builtin_templates, get_template_by_code,
)
from app.modules.expense.models import Expense, ExpenseBreakdown
from app.modules.auth.models import User, Department


# ===== 编号 =====
def _gen_form_no() -> str:
    return f"RB-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:4].upper()}"


# ===== 状态映射 =====
STATUS_LABELS = {
    "draft":      "草稿",
    "printed":    "已打印",
    "submitted":  "已提交",
    "reimbursed": "已报销",
    "done":       "已完成",
    "cancelled":  "已取消",
}


# ===== 模板（含自定义）=====
async def list_all_templates(db: AsyncSession) -> list[dict]:
    """内置 4 + DB 自定义，按 type 分组返回"""
    out = []
    for t in get_builtin_templates():
        out.append({
            "id": None,
            "code": t["code"],
            "name": t["name"],
            "type": t["type"],
            "icon": t.get("icon", "📋"),
            "color": t.get("color", "#4F6BFF"),
            "description": t.get("description"),
            "schema": t["schema"],
            "isSystem": True,
            "createdBy": None,
            "createdAt": None,
            "updatedAt": None,
        })
    rows = (await db.execute(select(ReimbursementTemplate).order_by(ReimbursementTemplate.id.asc()))).scalars().all()
    for r in rows:
        out.append({
            "id": r.id,
            "code": r.code,
            "name": r.name,
            "type": r.type,
            "icon": r.icon,
            "color": r.color,
            "description": r.description,
            "schema": r.schema_json,
            "isSystem": bool(r.is_system),
            "createdBy": r.created_by,
            "createdAt": r.created_at.isoformat() if r.created_at else None,
            "updatedAt": r.updated_at.isoformat() if r.updated_at else None,
        })
    return out


def list_templates_sync(db: AsyncSession | None = None) -> list[dict]:
    """同步 fallback：仅内置（router 内部使用）"""
    return get_builtin_templates()


async def get_template_by_code_async(db: AsyncSession, code: str) -> dict | None:
    """先用 code 查 DB 自定义，再退回内置"""
    r = (await db.execute(select(ReimbursementTemplate).where(ReimbursementTemplate.code == code))).scalars().first()
    if r:
        return {
            "id": r.id,
            "code": r.code,
            "name": r.name,
            "type": r.type,
            "icon": r.icon,
            "color": r.color,
            "description": r.description,
            "schema": r.schema_json,
            "isSystem": bool(r.is_system),
            "createdBy": r.created_by,
            "createdAt": r.created_at.isoformat() if r.created_at else None,
            "updatedAt": r.updated_at.isoformat() if r.updated_at else None,
        }
    return get_template_by_code(code)


# 保留旧同步函数避免破坏老调用
def list_templates() -> list[dict]:
    return get_builtin_templates()


def get_template(code: str) -> dict | None:
    return get_template_by_code(code)


# ===== 模板 CRUD =====
async def create_custom_template(db: AsyncSession, req, user_id: int) -> dict:
    code = (req.code or f"custom_{uuid.uuid4().hex[:8]}").strip()
    # 唯一性
    exist = (await db.execute(select(ReimbursementTemplate).where(ReimbursementTemplate.code == code))).scalars().first()
    if exist:
        raise ParamErrorException(f"模板 code 已存在：{code}")
    t = ReimbursementTemplate(
        code=code,
        name=req.name,
        type=req.type or "custom",
        icon=req.icon or "📋",
        color=req.color or "#4F6BFF",
        description=req.description,
        schema_json=req.schema,
        is_system=0,
        created_by=user_id,
    )
    db.add(t)
    await db.commit()
    await db.refresh(t)
    return {
        "id": t.id, "code": t.code, "name": t.name, "type": t.type,
        "icon": t.icon, "color": t.color, "description": t.description,
        "schema": t.schema_json, "isSystem": False,
        "createdBy": t.created_by,
        "createdAt": t.created_at.isoformat() if t.created_at else None,
        "updatedAt": t.updated_at.isoformat() if t.updated_at else None,
    }


async def update_custom_template(db: AsyncSession, req) -> dict:
    t = (await db.execute(select(ReimbursementTemplate).where(ReimbursementTemplate.id == req.templateId))).scalars().first()
    if not t:
        raise NotFoundException("模板不存在")
    if t.is_system:
        raise ParamErrorException("系统内置模板不可修改，请先复制为自定义模板")
    if req.name is not None: t.name = req.name
    if req.type is not None: t.type = req.type
    if req.icon is not None: t.icon = req.icon
    if req.color is not None: t.color = req.color
    if req.description is not None: t.description = req.description
    if req.schema is not None: t.schema_json = req.schema
    await db.commit()
    await db.refresh(t)
    return {
        "id": t.id, "code": t.code, "name": t.name, "type": t.type,
        "icon": t.icon, "color": t.color, "description": t.description,
        "schema": t.schema_json, "isSystem": bool(t.is_system),
        "createdBy": t.created_by,
        "createdAt": t.created_at.isoformat() if t.created_at else None,
        "updatedAt": t.updated_at.isoformat() if t.updated_at else None,
    }


async def delete_custom_template(db: AsyncSession, template_id: int) -> None:
    t = (await db.execute(select(ReimbursementTemplate).where(ReimbursementTemplate.id == template_id))).scalars().first()
    if not t:
        raise NotFoundException("模板不存在")
    if t.is_system:
        raise ParamErrorException("系统内置模板不可删除")
    # 检查是否被报销单引用
    used = (await db.execute(
        select(func.count(ReimbursementForm.id)).where(ReimbursementForm.template_type == t.code)
    )).scalar() or 0
    if used > 0:
        raise ParamErrorException(f"该模板已被 {used} 张报销单引用，无法删除")
    await db.delete(t)
    await db.commit()


async def clone_builtin_to_custom(db: AsyncSession, code: str, user_id: int, new_name: str | None = None) -> dict:
    """复制内置模板为自定义"""
    src = get_template_by_code(code)
    if not src:
        raise NotFoundException(f"内置模板不存在：{code}")
    new_code = f"custom_{code}_{uuid.uuid4().hex[:6]}"
    t = ReimbursementTemplate(
        code=new_code,
        name=new_name or (src["name"] + " · 副本"),
        type=src["type"],
        icon=src.get("icon", "📋"),
        color=src.get("color", "#4F6BFF"),
        description=src.get("description"),
        schema_json=src["schema"],
        is_system=0,
        created_by=user_id,
    )
    db.add(t)
    await db.commit()
    await db.refresh(t)
    return {
        "id": t.id, "code": t.code, "name": t.name, "type": t.type,
        "icon": t.icon, "color": t.color, "description": t.description,
        "schema": t.schema_json, "isSystem": False,
        "createdBy": t.created_by,
        "createdAt": t.created_at.isoformat() if t.created_at else None,
        "updatedAt": t.updated_at.isoformat() if t.updated_at else None,
    }



# ===== 列表 =====
async def list_forms(db: AsyncSession, page: int, page_size: int, keyword: str = "", filters: dict = None):
    filters = filters or {}
    q = select(ReimbursementForm)
    if keyword:
        like = f"%{keyword}%"
        q = q.where(or_(
            ReimbursementForm.form_no.ilike(like),
            ReimbursementForm.title.ilike(like),
        ))
    if filters.get("status") and filters["status"] != "all":
        q = q.where(ReimbursementForm.status == filters["status"])
    if filters.get("templateType") and filters["templateType"] != "all":
        q = q.where(ReimbursementForm.template_type == filters["templateType"])

    total = (await db.execute(select(func.count()).select_from(q.subquery()))).scalar() or 0
    q = q.options(
        selectinload(ReimbursementForm.applicant),
        selectinload(ReimbursementForm.department),
        selectinload(ReimbursementForm.details),
    ).order_by(ReimbursementForm.id.desc()).offset((page - 1) * page_size).limit(page_size)
    rows = (await db.execute(q)).scalars().all()

    items = [_form_to_dict(f) for f in rows]
    return items, total


# ===== 详情 =====
async def get_form(db: AsyncSession, form_id: int) -> dict:
    f = (await db.execute(
        select(ReimbursementForm)
        .where(ReimbursementForm.id == form_id)
        .options(
            selectinload(ReimbursementForm.applicant),
            selectinload(ReimbursementForm.department),
            selectinload(ReimbursementForm.details),
        )
    )).scalar_one_or_none()
    if not f:
        raise NotFoundException(f"报销单不存在：{form_id}")
    d = _form_to_dict(f, with_details=True)
    d["templateSnapshot"] = f.template_snapshot
    return d


# ===== 从销售费用生成报销单 =====
async def create_from_expenses(
    db: AsyncSession, req: ReimburseCreate, current_user_id: int,
) -> dict:
    if not req.expenseIds:
        raise ParamErrorException("请至少选择一条销售费用")

    # 加载选中的费用
    expenses = (await db.execute(
        select(Expense).where(Expense.id.in_(req.expenseIds))
        .options(selectinload(Expense.applicant))
    )).scalars().all()
    # 项目名懒加载（expense 模型无 project relationship）
    project_ids = list({e.project_id for e in expenses if e.project_id})
    project_map = {}
    if project_ids:
        from app.modules.project.models import Project
        rows = (await db.execute(select(Project.id, Project.name).where(Project.id.in_(project_ids)))).all()
        project_map = {r[0]: r[1] for r in rows}
    if len(expenses) != len(req.expenseIds):
        raise ParamErrorException("部分费用记录不存在或无权限")

    # 申请人默认取当前用户
    applicant = (await db.execute(select(User).where(User.id == current_user_id))).scalar_one()
    department = None
    if applicant.department_id:
        department = (await db.execute(select(Department).where(Department.id == applicant.department_id))).scalar_one_or_none()

    # 计算总金额（分）
    total = sum([int(e.amount or 0) for e in expenses])
    title = req.title or f"{applicant.name}的{_template_label(req.templateType)}"

    form = ReimbursementForm(
        form_no=_gen_form_no(),
        template_type=req.templateType or "general",
        title=title,
        applicant_id=current_user_id,
        department_id=department.id if department else None,
        total_amount=total,
        actual_amount=0,
        status="draft",
        expense_date=req.expenseDate or date.today(),
        remark=req.remark,
        ai_description=req.aiDescription,
        template_snapshot=get_template_by_code(req.templateType) or get_template_by_code("general"),
    )
    db.add(form)
    await db.flush()

    for seq, e in enumerate(expenses, 1):
        project_name = project_map.get(e.project_id) if e.project_id else None
        client_name = _extract_client_name(e.title or "")
        det = ReimbursementDetail(
            form_id=form.id,
            expense_id=e.id,
            expense_code=e.code,
            expense_type=e.category,
            expense_date=e.expense_date,
            client_name=client_name,
            project_name=project_name,
            title=e.title,
            amount=int(e.amount or 0),
            seq=seq,
        )
        db.add(det)

    await db.commit()
    await db.refresh(form)
    return await get_form(db, form.id)


def _template_label(t: str) -> str:
    return {
        "general": "通用费用报销",
        "travel": "差旅费报销",
        "hospitality": "业务招待费报销",
        "marketing": "市场推广费报销",
        "custom": "自定义报销",
    }.get(t, "报销")


def _extract_client_name(text: str) -> str | None:
    """从费用 title 尝试解析客户名（启发式：'拜访XX' / '接待XX' / 'XX推广' 等）"""
    if not text:
        return None
    for pat in [r"拜访([\u4e00-\u9fa5A-Za-z0-9]{2,12})", r"接待([\u4e00-\u9fa5A-Za-z0-9]{2,12})",
                r"([\u4e00-\u9fa5A-Za-z0-9]{2,12})推广", r"([\u4e00-\u9fa5A-Za-z0-9]{2,12})差旅"]:
        m = re.search(pat, text)
        if m:
            return m.group(1)
    return None


# ===== 更新 =====
async def update_form(db: AsyncSession, req: ReimburseUpdate) -> dict:
    f = (await db.execute(
        select(ReimbursementForm).where(ReimbursementForm.id == req.formId)
    )).scalar_one_or_none()
    if not f:
        raise NotFoundException(f"报销单不存在：{req.formId}")
    if f.status not in ("draft", "printed"):
        raise ParamErrorException(f"当前状态「{STATUS_LABELS.get(f.status, f.status)}」不可编辑")

    if req.title is not None:
        f.title = req.title
    if req.templateType:
        f.template_type = req.templateType
        f.template_snapshot = get_template_by_code(req.templateType) or f.template_snapshot
    if req.expenseDate:
        f.expense_date = req.expenseDate
    if req.remark is not None:
        f.remark = req.remark
    if req.aiDescription is not None:
        f.ai_description = req.aiDescription

    if req.expenseIds is not None:
        # 重新加载明细
        old_details = (await db.execute(
            select(ReimbursementDetail).where(ReimbursementDetail.form_id == f.id)
        )).scalars().all()
        for d in old_details:
            await db.delete(d)
        await db.flush()

        expenses = (await db.execute(
            select(Expense).where(Expense.id.in_(req.expenseIds))
        )).scalars().all()
        # 项目名懒加载
        proj_ids = list({e.project_id for e in expenses if e.project_id})
        proj_map = {}
        if proj_ids:
            from app.modules.project.models import Project
            rows2 = (await db.execute(select(Project.id, Project.name).where(Project.id.in_(proj_ids)))).all()
            proj_map = {r[0]: r[1] for r in rows2}
        for seq, e in enumerate(expenses, 1):
            project_name = proj_map.get(e.project_id) if e.project_id else None
            det = ReimbursementDetail(
                form_id=f.id,
                expense_id=e.id,
                expense_code=e.code,
                expense_type=e.category,
                expense_date=e.expense_date,
                client_name=_extract_client_name(e.title or ""),
                project_name=project_name,
                title=e.title,
                amount=int(e.amount or 0),
                seq=seq,
            )
            db.add(det)
        f.total_amount = sum([int(e.amount or 0) for e in expenses])
        f.actual_amount = 0  # 重置

    await db.commit()
    return await get_form(db, f.id)


# ===== 删除 =====
async def delete_form(db: AsyncSession, form_id: int):
    f = (await db.execute(
        select(ReimbursementForm)
        .where(ReimbursementForm.id == form_id)
        .options(selectinload(ReimbursementForm.details))
    )).scalar_one_or_none()
    if not f:
        raise NotFoundException(f"报销单不存在：{form_id}")
    # 删除时回退关联销售费用状态：paid → submitted
    from app.modules.expense.models import Expense
    for det in f.details or []:
        if det.expense_id:
            exp = (await db.execute(select(Expense).where(Expense.id == det.expense_id))).scalar_one_or_none()
            if exp and exp.status == "paid":
                exp.status = "submitted"
    # 先删明细
    for det in f.details or []:
        await db.delete(det)
    await db.delete(f)
    await db.commit()


# ===== 标记已打印 =====
async def mark_printed(db: AsyncSession, form_id: int) -> dict:
    f = (await db.execute(select(ReimbursementForm).where(ReimbursementForm.id == form_id))).scalar_one_or_none()
    if not f:
        raise NotFoundException(f"报销单不存在：{form_id}")
    if f.status == "draft":
        f.status = "printed"
        await db.commit()
        await db.refresh(f)
    return await get_form(db, f.id)


# ===== 回填实际报销 → 自动回写销售费用 =====
async def fillback(db: AsyncSession, req: ReimburseFillback) -> dict:
    f = (await db.execute(
        select(ReimbursementForm).where(ReimbursementForm.id == req.formId)
        .options(selectinload(ReimbursementForm.details))
    )).scalar_one_or_none()
    if not f:
        raise NotFoundException(f"报销单不存在：{req.formId}")
    if req.actualAmount < 0:
        raise ParamErrorException("实际金额不能为负")
    if f.status == "cancelled":
        raise ParamErrorException("已取消的报销单不可回填")

    f.actual_amount = req.actualAmount
    if req.paymentDate:
        f.payment_date = req.paymentDate
    if req.voucherNo:
        f.voucher_no = req.voucherNo
    if req.remark is not None:
        f.remark = req.remark

    # 把每个明细回填金额写到关联的销售费用上
    detail_amounts = req.detailAmounts or {}
    # 先把所有明细的 reimbursed_amount 写到对象（未 flush）
    for det in f.details:
        per = int(detail_amounts.get(str(det.expense_id), det.amount or 0))
        det.reimbursed_amount = per
    await db.flush()  # 立即 flush 让后续 SUM 看到本次写

    # 找出本次涉及的所有 expense_id
    affected_expense_ids = list({det.expense_id for det in f.details})
    if affected_expense_ids:
        # 一次性 SUM 算每个 expense 的累计已报销金额
        sum_rows = (await db.execute(
            select(ReimbursementDetail.expense_id, func.coalesce(func.sum(ReimbursementDetail.reimbursed_amount), 0))
            .where(ReimbursementDetail.expense_id.in_(affected_expense_ids))
            .group_by(ReimbursementDetail.expense_id)
        )).all()
        cum_map = {r[0]: int(r[1]) for r in sum_rows}
        # 加载这些 expense
        exps = (await db.execute(select(Expense).where(Expense.id.in_(affected_expense_ids)))).scalars().all()
        for exp in exps:
            cum = cum_map.get(exp.id, 0)
            exp_total = int(exp.amount or 0)
            if cum >= exp_total and exp_total > 0:
                exp.status = "paid"
                exp.finish_at = datetime.utcnow()
            elif cum > 0:
                exp.status = "approved"
                exp.finish_at = None

    f.status = "reimbursed" if req.actualAmount > 0 else f.status
    f.status = "done"  # 直接完成（用户要求：不走审批流）

    await db.commit()
    return await get_form(db, f.id)


# ===== AI 生成说明 =====
def ai_generate_description(expense_ids: list[int], expenses_data: list[dict] | None = None) -> str:
    """根据费用数据启发式生成报销说明（轻量规则，不调外部 AI）"""
    if not expenses_data and not expense_ids:
        return "因业务活动产生相关费用，现申请报销。"

    items = expenses_data or []
    if not items:
        return f"因业务活动产生相关费用（{len(expense_ids)} 笔），现申请报销。"

    cats = set([(e.get("category") or "") for e in items if e.get("category")])
    clients = set([(e.get("clientName") or "") for e in items if e.get("clientName")])
    total = sum([int(e.get("amount") or 0) for e in items]) / 100.0

    parts = []
    if "travel" in cats or "差旅" in cats:
        parts.append("差旅活动")
    if "hospitality" in cats or "招待" in cats:
        parts.append("客户拜访与业务招待")
    if "marketing" in cats or "推广" in cats:
        parts.append("市场推广与广告投放")
    if "office" in cats or "办公" in cats:
        parts.append("日常办公采购")
    if "training" in cats or "培训" in cats:
        parts.append("员工培训")
    if not parts:
        parts.append("业务活动")

    main = parts[0] if len(parts) == 1 else "、".join(parts[:-1]) + "及" + parts[-1]

    client_part = ""
    if clients:
        c = list(clients)[0]
        client_part = f"（涉及{c}等客户）"

    return f"因{main}{client_part}产生相关费用，共 {len(items)} 笔，合计 ¥{total:,.2f}，现申请报销。恳请财务审核。"


# ===== AI 风险检测 =====
async def ai_risk_check(db: AsyncSession, expense_ids: list[int], form_id: int | None = None) -> dict:
    """重复报销 / 金额异常 / 超预算 风险检测"""
    if not expense_ids:
        return {"level": "low", "reasons": []}

    reasons = []
    # 1. 重复检测：在报销明细中查找同金额 / 同日期 / 同客户
    expenses = (await db.execute(
        select(Expense).where(Expense.id.in_(expense_ids))
    )).scalars().all()

    for e in expenses:
        # 同金额 + 同日期 + 状态已 paid
        dup = (await db.execute(
            select(func.count(Expense.id))
            .where(
                Expense.id != e.id,
                Expense.amount == e.amount,
                Expense.expense_date == e.expense_date,
                Expense.status == "paid",
            )
        )).scalar() or 0
        if dup > 0:
            reasons.append(f"费用「{e.title}」（¥{int(e.amount)/100:.2f}）存在已报销的同金额/同日期记录，请确认是否重复报销")

    # 2. 金额异常
    big = [e for e in expenses if int(e.amount or 0) >= 1000000]  # ≥ 1 万
    for e in big:
        reasons.append(f"费用「{e.title}」单笔金额 ¥{int(e.amount)/100:.2f} 较高，建议附加明细票据")

    # 3. 跨月报销
    today = date.today()
    for e in expenses:
        if e.expense_date and (today - e.expense_date).days > 90:
            reasons.append(f"费用「{e.title}」发生于 {e.expense_date}，距今超过 90 天，请确认是否超期")

    # 风险等级
    if any(["重复" in r for r in reasons]):
        level = "high"
    elif len(reasons) >= 2:
        level = "medium"
    elif reasons:
        level = "low"
    else:
        level = "none"

    return {"level": level, "reasons": reasons}


# ===== 工具 =====
def _form_to_dict(f: ReimbursementForm, with_details: bool = False) -> dict:
    d = {
        "formId": f.id,
        "id": f.id,
        "formNo": f.form_no,
        "templateType": f.template_type,
        "title": f.title,
        "applicant":   {"userId": f.applicant.id, "name": f.applicant.name} if f.applicant else None,
        "applicantId": f.applicant_id,
        "applicantName": f.applicant.name if f.applicant else None,
        "department":  {"id": f.department.id, "name": f.department.name} if f.department else None,
        "departmentName": f.department.name if f.department else None,
        "totalAmount": int(f.total_amount or 0),
        "actualAmount": int(f.actual_amount or 0),
        "currency": f.currency or "CNY",
        "status": f.status,
        "statusLabel": STATUS_LABELS.get(f.status, f.status),
        "expenseDate": f.expense_date,
        "paymentDate": f.payment_date,
        "voucherNo": f.voucher_no,
        "aiDescription": f.ai_description,
        "aiRiskFlag": f.ai_risk_flag,
        "aiRiskReason": f.ai_risk_reason,
        "remark": f.remark,
        "createdAt": f.created_at,
        "updatedAt": f.updated_at,
        "detailCount": len(f.details or []),
    }
    if with_details:
        d["details"] = [
            {
                "id": det.id,
                "expenseId": det.expense_id,
                "expenseCode": det.expense_code,
                "expenseType": det.expense_type,
                "expenseDate": det.expense_date,
                "clientName": det.client_name,
                "projectName": det.project_name,
                "title": det.title,
                "amount": int(det.amount or 0),
                "reimbursedAmount": int(det.reimbursed_amount or 0),
                "remark": det.remark,
                "seq": det.seq,
            }
            for det in sorted(f.details or [], key=lambda x: x.seq)
        ]
    return d

# ===== 模板识别（启发式）=====
import json as _json

_FIELD_HINTS = [
    # (key, label, regex/keywords) — 顺序影响优先级
    ("applicant",      "报销人",   [r"报销人[：:]?", r"经办人", r"出差人", r"申请人", r"报销人签字", r"出差人姓名"]),
    ("approver",       "审批人",   [r"审批人", r"核准[：:]?", r"总经理审批", r"领导审批"]),
    ("accountant",     "会计",     [r"会计[：:]?", r"财务审核", r"出纳[：:]?", r"复核[：:]?"]),
    ("department",     "部门",     [r"部门", r"所属部门", r"费用归属", r"报销部门"]),
    ("expenseDate",    "费用日期", [r"报销日期", r"费用日期", r"出差日期", r"发生日期", r"日期", r"年\s*月\s*日"]),
    ("formNo",         "单据编号", [r"单据编号", r"报销单号", r"编号", r"No", r"NO"]),
    ("summary",        "费用摘要", [r"费用摘要", r"摘要", r"事项", r"用途", r"事由"]),
    ("bigAmount",      "金额大写", [r"金额（大写）", r"金额大写", r"大写金额", r"人民币大写"]),
    ("smallAmount",    "金额小写", [r"金额（小写）", r"金额小写", r"小写金额", r"￥", r"¥"]),
    ("totalAmount",    "申请金额", [r"报销金额", r"申请金额", r"合计", r"金额合计", r"总额", r"差旅费合计"]),
    ("actualAmount",   "实报金额", [r"实报金额", r"实际报销", r"实付金额", r"实际支付"]),
    ("voucherNo",      "凭证号",   [r"凭证号", r"凭证编号", r"记账凭证"]),
    ("paymentDate",    "支付日期", [r"支付日期", r"付款日期"]),
    ("projectName",    "项目",     [r"项目", r"项目名称", r"工程项目"]),
    ("clientName",     "客户/供应商", [r"客户", r"供应商", r"招待对象", r"往来单位"]),
    ("receiptCount",   "附单据张数", [r"附单据张数", r"票据张数", r"附件张数"]),
    ("remark",         "备注",     [r"备注", r"说明"]),
]
_SIG_HINTS = [r"报销人签字", r"部门负责人", r"财务复核", r"财务审核", r"总经理审批", r"审批人", r"出纳", r"经办人签字", r"领导审批", r"项目负责人", r"项目经理", r"核准", r"会计", r"报销人", r"审批"]


def _classify_field(name: str) -> str:
    """粗略分类字段类型"""
    if any(k in name for k in ["金额", "总额", "合计", "费用"]):
        return "money"
    if any(k in name for k in ["日期", "时间"]):
        return "date"
    if any(k in name for k in ["数量", "单价", "税率"]):
        return "number"
    return "text"



def extract_text_from_bytes(raw: bytes, filename: str = "") -> str:
    """从多种文件格式中提取文本（xlsx 优先，文本格式 fallback）"""
    name = (filename or "").lower()
    # 1) .xlsx 用 openpyxl 读所有 sheet 所有单元格
    if name.endswith(".xlsx") or raw[:2] == b"PK":
        try:
            import io
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
            lines = []
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
                    if cells:
                        lines.append("\t".join(cells))
            return "\n".join(lines)
        except Exception as e:
            # 退化：尝试作 zip 解 sharedStrings
            try:
                import io, zipfile, re
                z = zipfile.ZipFile(io.BytesIO(raw))
                xml = z.read("xl/sharedStrings.xml").decode("utf-8", errors="ignore")
                strs = re.findall(r"<t[^>]*>([^<]*)</t>", xml)
                return "\n".join(strs)
            except Exception:
                return ""
    # 2) 文本格式
    for enc in ("utf-8", "utf-8-sig", "gbk", "gb2312", "latin-1"):
        try:
            return raw.decode(enc)
        except Exception:
            continue
    return raw.decode("utf-8", errors="ignore")





def extract_xlsx_table(raw: bytes) -> dict | None:
    """从 .xlsx 提取真实表格结构（包含合并单元格信息）
    返回: {"sheet": "Sheet1", "cols": 7, "rows": [[{text, colspan, rowspan, type?, key?}, ...], ...]}
    如果解析失败返回 None
    """
    try:
        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=False)
        # 选第一个非空 sheet
        ws = None
        for s in wb.worksheets:
            if s.max_row > 0 and s.max_column > 0:
                ws = s
                break
        if ws is None:
            return None
        # 1) 算实际最大非空列（去尾部空白列）
        max_col_used = 0
        for row in ws.iter_rows(values_only=False):
            for c in row:
                if c.value is not None and c.column > max_col_used:
                    max_col_used = c.column
        # 2) 算实际最大非空行
        max_row_used = 0
        for row in ws.iter_rows(values_only=False):
            for c in row:
                if c.value is not None and c.row > max_row_used:
                    max_row_used = c.row
        # 3) 合并信息：只保留与 max_col_used/max_row_used 相交的合并
        merged_map = {}  # (row, col) -> (max_row, max_col)
        for r in ws.merged_cells.ranges:
            if r.max_row > max_row_used or r.max_col > max_col_used:
                continue
            for rr in range(r.min_row, r.max_row + 1):
                for cc in range(r.min_col, r.max_col + 1):
                    merged_map[(rr, cc)] = (r.max_row, r.max_col)
        # 4) 逐行输出：只输出每个合并区域的左上角 cell + 非合并区 cell
        #    被合并区"覆盖"的位置不输出（HTML table 中自动被合并区占据）
        out_rows = []
        for r in range(1, max_row_used + 1):
            row_cells = []
            c = 1
            while c <= max_col_used:
                # 如果当前位置 (r,c) 被某个合并区"内部覆盖"（不是左上角），跳过
                top_left = merged_map.get((r, c))
                if top_left and (top_left != (r, c)):
                    # 找这个 (r,c) 所属的合并区左上角
                    pass
                cell = ws.cell(row=r, column=c)
                v = cell.value
                if (r, c) in merged_map:
                    max_r, max_c = merged_map[(r, c)]
                    # 只在左上角 (r==min_row, c==min_col) 输出
                    # 简单判断：merged_map[(r,c)] == (r,c) 时是左上角？
                    # 不对，需要从合并区 ranges 找
                    # 用另一个 map：start_map[(min_row, min_col)] = (max_row, max_col)
                    pass
                # 用更可靠的方式：track 哪些 cell 是合并区左上角
                c += 1
            out_rows.append(row_cells)
        # 重写：维护 is_top_left_map
        is_top_left = set()  # (row, col) 是某个合并区的左上角
        starts = {}  # (min_row, min_col) -> (max_row, max_col)
        for r in ws.merged_cells.ranges:
            if r.max_row > max_row_used or r.max_col > max_col_used:
                continue
            is_top_left.add((r.min_row, r.min_col))
            starts[(r.min_row, r.min_col)] = (r.max_row, r.max_col)
        out_rows = []
        for r in range(1, max_row_used + 1):
            row_cells = []
            c = 1
            while c <= max_col_used:
                if (r, c) in is_top_left:
                    max_r, max_c = starts[(r, c)]
                    v = ws.cell(row=r, column=c).value
                    text = (str(v).strip() if v is not None else "")
                    row_cells.append({"text": text, "colspan": max_c - c + 1, "rowspan": max_r - r + 1, "isHeader": r == 1})
                    c = max_c + 1
                elif (r, c) in merged_map:
                    # 被合并区内部覆盖，跳过
                    c += 1
                else:
                    v = ws.cell(row=r, column=c).value
                    text = (str(v).strip() if v is not None else "")
                    if text or True:  # 输出以保持列对齐
                        row_cells.append({"text": text, "colspan": 1, "rowspan": 1, "isHeader": r == 1})
                    c += 1
            out_rows.append(row_cells)
        return {"sheet": ws.title, "cols": max_col_used, "rows": out_rows}
    except Exception as e:
        return None


def _infer_field_type(text: str) -> str:
    """从 cell text 推断字段类型（用于单笔模板字段标记）"""
    t = text or ""
    if "金额" in t and "大写" in t: return "bigAmount"
    if "金额" in t and "小写" in t: return "smallAmount"
    if "金额" in t: return "money"
    if "日期" in t or "年" in t: return "date"
    if "数量" in t or "张数" in t: return "number"
    if "税率" in t: return "number"
    return "text"


def build_schema_from_table(table: dict) -> dict:
    """从 xlsx table 反推 schema（1:1 还原）
    - 第 0 行 = 标题行（合并跨 ≥ 3 列），保留在 table.rows[0]
    - 含"核准/审批/会计/报销人"且为多 cell 短文本行 = 签字栏 labelOnly
    - 其它 label+value 行 = summary.fields
    - 输出 layout=single + table + fields + signatures + labelOnly
    - 保留所有行（含合并区内部空 cell 占位行），由前端用 rowspan 渲染
    """
    import sys
    rows_all = list(table.get("rows", []))
    cols = table.get("cols", 7)
    sys.stderr.write(f"[D] INPUT rows={len(rows_all)} cols={cols}\n"); sys.stderr.flush()
    rows = rows_all

    # 1) 检测标题行（合并跨 ≥ 3 列），但不删除 — 保留到 table.rows[0]
    title = "费用报销单"
    if rows and len(rows[0]) >= 1:
        first = rows[0][0]
        if first.get("colspan", 1) >= max(3, cols - 1):
            title = first["text"] or title
            sys.stderr.write(f"[D] TITLE detected: {title!r}, rows unchanged\n"); sys.stderr.flush()

    # 1.5) 跳过日期/编号等单 cell 行（含"年/月/日/编号/No"等特征）
    # 这些通常是模板 1 标题下的"年 月 日"行
    filtered = []
    for row in rows:
        if len(row) == 1 and row[0].get("colspan", 1) >= 3:
            text = row[0].get("text", "")
            if any(k in text for k in ["年", "月", "日", "编号", "No", "NO"]):
                continue
        filtered.append(row)
    rows = filtered
    sys.stderr.write(f"[D] AFTER 1.5: rows={len(rows)}\n"); sys.stderr.flush()

    # 1.6) 去重：xlsx 多个相同模板时只取第一份
    if len(rows) > 8:
        # 找第一个签字行（多等分 label-only 短文本行）
        def _is_sig_row(r):
            if len(r) < 2: return False
            if not all(c.get("colspan", 1) == 1 for c in r): return False
            labels = [c["text"].rstrip("：:") for c in r]
            return sum(1 for t in labels if t in ("核准", "审批", "会计", "报销人", "出纳", "财务")) >= 2
        first_sig = None
        for idx, r in enumerate(rows):
            if _is_sig_row(r):
                first_sig = idx
                break
        if first_sig is not None:
            # 模板 1 范围 = rows[0..first_sig]（含签字行）
            # 跳过所有空 cell 行（cs=0 或全部空 text）直到下一个有内容的行
            end = first_sig + 1
            while end < len(rows) and (not rows[end] or all(not c.get("text", "").strip() for c in rows[end])):
                end += 1
            # end 是模板 2 起始
            if end < len(rows) - 3:
                sys.stderr.write(f"[D] DEDUP by sig: first_sig={first_sig}, end={end}, rows -> {end}\n"); sys.stderr.flush()
                rows = rows[:end]
                # 清理尾部空行（cs=0 或全部空 text）
                while rows and (not rows[-1] or all(not c.get("text", "").strip() for c in rows[-1])):
                    rows.pop()
                sys.stderr.write(f"[D] After trim trailing empty: rows={len(rows)}\n"); sys.stderr.flush()
            else:
                sys.stderr.write(f"[D] NO DEDUP needed (no second template)\n"); sys.stderr.flush()
        else:
            sys.stderr.write(f"[D] NO SIG ROW found\n"); sys.stderr.flush()

    # 2) 推断签字栏（从所有行找，含标题行也可能误判，所以限定 colspan=1 短文本）
    sig_labels = []
    body_rows = []
    for row in rows:
        texts = [c["text"].rstrip("：:") for c in row]
        sig_like = [t for t in texts if t in ("核准", "审批", "会计", "报销人", "出纳", "财务", "负责人", "经办人")]
        if len(sig_like) >= 2 and len(row) >= 2 and all(c.get("colspan", 1) == 1 for c in row):
            sig_labels = sig_like
        else:
            body_rows.append(row)

    # 3) 字段网格（跳过标题行 — 用 isHeader 标记）
    fields = []
    for row in body_rows:
        if not row: continue
        non_empty = [c for c in row if c.get("text", "").strip()]
        if not non_empty: continue
        if all(c.get("isHeader") for c in row): continue
        if len(row) == 1:
            label_cell = row[0]
            label = label_cell["text"].strip()
            if label and any(k in label for k in ["摘要", "金额", "备注", "张数", "凭证", "项目", "客户", "部门", "报销人", "日期"]):
                fields.append({"key": _infer_field_type(label), "label": label, "type": _infer_field_type(label)})
            continue
        label_cell = row[0]
        value_cell = row[1] if len(row) > 1 else None
        label = label_cell["text"].strip()
        value = value_cell["text"].strip() if value_cell else ""
        if label:
            if len(row) >= 4 and row[2].get("text", "").strip():
                fields.append({"key": _infer_field_type(label), "label": label, "type": _infer_field_type(label), "value": value})
                l2 = row[2]["text"].strip()
                v2 = row[3]["text"].strip() if len(row) > 3 else ""
                fields.append({"key": _infer_field_type(l2), "label": l2, "type": _infer_field_type(l2), "value": v2})
            else:
                fields.append({"key": _infer_field_type(label), "label": label, "type": _infer_field_type(label), "value": value})

    sigs = [{"key": re.sub(r"[\\s\\W]+", "_", s) or "sig", "label": s} for s in sig_labels] or [
        {"key": "applicant", "label": "报销人"},
        {"key": "dept_lead", "label": "部门负责人"},
        {"key": "finance",   "label": "财务审核"},
    ]

    # 4) 保留所有行（含合并区内部空 cell 占位行 []），由前端 normalize + rowspan 渲染
    # 关键：合并区的"内部"行（colspan 起点 cell 所在行的后续 1..rowspan-1 行）必须保留为 []
    # 否则 HTML table 中 rowspan 起点 cell 会试图占据不存在的行
    final_rows = [r for r in rows if r is not None]
    rows = final_rows
    sys.stderr.write(f"[D] After clean None rows: rows={len(rows)}\n"); sys.stderr.flush()

    result = {
        "layout": "single",
        "header": {
            "company": "上海数智信息技术有限公司",
            "title": title,
            "subtitle": "Reimbursement Form",
        },
        "applicant": {"fields": []},
        "summary": {"fields": fields},
        "details": {"columns": []},
        "table": {
            "cols": cols,
            "rows": rows,
        },
        "footer": {
            "signatures": sigs,
            "labelOnly": True,
            "note": "本报销单一式两份，财务与报销人各执一份。",
        },
    }
    sys.stderr.write(f"[D] OUTPUT table.rows={len(rows)} title={title!r} fields={len(fields)} sigs={len(sigs)}\n"); sys.stderr.flush()
    return result


def recognize_template_from_text(text: str) -> dict:
    """从粘文本识别 schema（启发式）"""
    if not text or not text.strip():
        return {"detectedFields": [], "suggestedSchema": _empty_schema("识别结果"), "textPreview": "", "confidence": 0.0}

    text_clean = text.replace("\r\n", "\n").replace("\r", "\n")
    lines = [l.strip() for l in text_clean.split("\n") if l.strip()]
    full = "\n".join(lines)

    detected = []
    seen = set()
    for key, label, patterns in _FIELD_HINTS:
        for p in patterns:
            if re.search(p, full):
                if key not in seen:
                    detected.append({
                        "key": key,
                        "label": label,
                        "type": _classify_field(label),
                    })
                    seen.add(key)
                break

    # 签字栏
    sigs = []
    seen_sig_labels = set()
    for p in _SIG_HINTS:
        if re.search(p, full):
            label_clean = re.sub(r"[：:]+$", "", p)
            if label_clean in seen_sig_labels:
                continue
            seen_sig_labels.add(label_clean)
            sigs.append({"label": label_clean, "key": re.sub(r"[\\s\\W]+", "_", label_clean) or "sig"})

    # 推断 title（按用户文本特征匹配）
    title = "费用报销单"
    for kw, guess in [("研发", "研发费用报销单"), ("差旅", "差旅费报销单"), ("招待", "业务招待费报销单"), ("市场", "市场推广费报销单"), ("通用", "通用费用报销单"), ("项目", "项目费用报销单")]:
        if kw in full[:200]:
            title = guess
            break

    # 推断 type
    t = "custom"
    for k, v in [("研发", "rd"), ("差旅", "travel"), ("招待", "hospitality"), ("市场", "marketing"), ("通用", "general"), ("项目", "project")]:
        if k in full[:200]:
            t = v
            break

    # 推断 layout：含"附单据张数/金额大写/金额小写/费用摘要"等"单笔"特征 → single 模式
    # 否则默认 detail 模式（多笔费用明细表）
    single_keys = ("summary", "bigAmount", "smallAmount", "receiptCount", "remark")
    is_single = any(f["key"] in single_keys for f in detected) or any(
        kw in full for kw in ["附单据张数", "金额（大写）", "金额大写", "金额（小写）", "金额小写", "费用摘要"]
    )

    # applicant 区
    applicant_fields = [f for f in detected if f["key"] in ("applicant", "department", "expenseDate", "formNo")][:4]

    if is_single:
        # 单笔：所有非 applicant 字段进 summary（label + value 网格），无 details
        # 排除签名类（approver/accountant 已在 signatures 出现）
        single_fields = [f for f in detected if f["key"] not in ("applicant", "department", "expenseDate", "formNo", "approver", "accountant")]
        # 排序：金额大写 → 金额小写 → 总金额 → 实报 → 摘要 → 凭证号 → 附单据张数 → 备注
        order = {"summary": 1, "bigAmount": 2, "smallAmount": 3, "totalAmount": 4, "actualAmount": 5, "voucherNo": 6, "receiptCount": 7, "remark": 8, "clientName": 9, "projectName": 10, "paymentDate": 11}
        single_fields.sort(key=lambda f: order.get(f["key"], 99))
        # 真实样张的"核准/审批/会计/报销人"风格是标签横排（无横线）
        sigs_single = sigs or []
        suggested = {
            "layout": "single",
            "header": {
                "company": "上海数智信息技术有限公司",
                "title": title,
                "subtitle": "Reimbursement Form",
            },
            "applicant": {"fields": applicant_fields},
            "summary": {"fields": single_fields},
            "details": {"columns": []},
            "footer": {
                "signatures": sigs_single,
                "labelOnly": True,  # single 模式：签字栏为 label 横排，无签字线
                "note": "本报销单一式两份，财务与报销人各执一份。",
            },
        }
    else:
        # 多笔：保留明细表
        summary_fields = [f for f in detected if f["key"] in ("totalAmount", "actualAmount", "voucherNo")]
        suggested = {
            "layout": "detail",
            "header": {
                "company": "上海数智信息技术有限公司",
                "title": title,
                "subtitle": "Reimbursement Form",
            },
            "applicant": {"fields": applicant_fields},
            "summary": {"fields": summary_fields},
            "details": {
                "columns": [
                    {"key": "seq",         "label": "#",          "width": 32,  "align": "center"},
                    {"key": "expenseDate", "label": "费用日期",   "width": 90,  "align": "center"},
                    {"key": "expenseType", "label": "费用类型",   "width": 80,  "align": "center"},
                    {"key": "title",       "label": "摘要",       "width": 200, "align": "left"},
                    {"key": "clientName",  "label": "客户/供应商", "width": 130, "align": "left"},
                    {"key": "amount",      "label": "金额(元)",   "width": 100, "align": "right", "type": "money"},
                ]
            },
            "footer": {
                "signatures": sigs or [
                    {"label": "报销人签字", "key": "applicant"},
                    {"label": "部门负责人", "key": "dept_lead"},
                    {"label": "财务审核",   "key": "finance"},
                ],
                "note": "本报销单一式两份，财务与报销人各执一份。",
            },
        }

    confidence = min(1.0, 0.2 + 0.1 * len(detected) + 0.05 * len(sigs))

    return {
        "detectedFields": detected,
        "suggestedSchema": suggested,
        "textPreview": full[:600],
        "confidence": round(confidence, 2),
    }


def _empty_schema(title: str = "自定义报销单") -> dict:
    return {
        "header": {"company": "上海数智信息技术有限公司", "title": title, "subtitle": "Reimbursement Form"},
        "applicant": {"fields": [
            {"key": "applicant",   "label": "报销人"},
            {"key": "department",  "label": "部门"},
            {"key": "expenseDate", "label": "费用日期"},
            {"key": "formNo",      "label": "单据编号"},
        ]},
        "summary": {"fields": [
            {"key": "totalAmount",  "label": "申请金额", "type": "money"},
            {"key": "actualAmount", "label": "实报金额", "type": "money"},
            {"key": "voucherNo",    "label": "凭证号"},
        ]},
        "details": {"columns": [
            {"key": "seq",         "label": "#",          "width": 32,  "align": "center"},
            {"key": "expenseDate", "label": "费用日期",   "width": 90,  "align": "center"},
            {"key": "expenseType", "label": "费用类型",   "width": 80,  "align": "center"},
            {"key": "title",       "label": "摘要",       "width": 200, "align": "left"},
            {"key": "clientName",  "label": "客户/供应商", "width": 130, "align": "left"},
            {"key": "amount",      "label": "金额(元)",   "width": 100, "align": "right", "type": "money"},
        ]},
        "footer": {"signatures": [
            {"label": "报销人签字", "key": "applicant"},
            {"label": "部门负责人", "key": "dept_lead"},
            {"label": "财务审核",   "key": "finance"},
        ], "note": "本报销单一式两份，财务与报销人各执一份。"},
    }



async def batch_delete_forms(db: AsyncSession, form_ids: list[int], operator_id: int) -> dict:
    """批量删除报销单（仅允许草稿/已取消状态，已打印/已完成的跳过）"""
    from app.modules.reimbursement.models import ReimbursementForm
    rows = (await db.execute(
        select(ReimbursementForm).where(ReimbursementForm.id.in_(form_ids))
    )).scalars().all()
    deleted = 0
    skipped = []
    for f in rows:
        if f.status in ("done", "printed", "reimbursed"):
            skipped.append({"formId": f.id, "formNo": f.form_no, "reason": f"状态 {f.status} 不可删除"})
            continue
        await db.delete(f)
        deleted += 1
    await db.commit()
    return {"deleted": deleted, "skipped": skipped}